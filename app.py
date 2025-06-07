from datetime import datetime
import math
from flask import Flask, request, jsonify, send_from_directory
from flask_cors import CORS
import os
import random
from flask_limiter import Limiter
from flask_limiter.util import get_remote_address
import pandas as pd
import numpy as np
from functools import wraps
from dotenv import load_dotenv
import os

from openpyxl.styles import *
from openpyxl.utils import get_column_letter

from werkzeug.utils import secure_filename

app = Flask(__name__, static_folder='frontend', static_url_path='')
CORS(app)

limiter = Limiter(
    get_remote_address,
    app=app,
    default_limits=["10 per minute"]
)

# Configuración de carpetas
UPLOAD_FOLDER = os.getenv('PATH_UPLOAD', 'uploads')
PROCESSED_FOLDER = os.getenv('PATH_PROCESSED', 'static/processed')
ALLOWED_EXTENSIONS = {'xlsx'}

app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['PROCESSED_FOLDER'] = PROCESSED_FOLDER

load_dotenv()

AUTH_USERNAME = os.getenv('BASIC_AUTH_USERNAME')
AUTH_PASSWORD = os.getenv('BASIC_AUTH_PASSWORD')

# Crear carpetas si no existen
os.makedirs(UPLOAD_FOLDER, exist_ok=True)
os.makedirs(PROCESSED_FOLDER, exist_ok=True)


def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


def generate_color():
    r = random.randint(150, 200)
    g = random.randint(150, 200)
    b = random.randint(0, 200)
    
    color_hex = "{:02X}{:02X}{:02X}".format(r, g, b)
    
    return color_hex


def set_styless(book):
    font = Font(name='Calibri', size=11)
    font_bold = Font(name='Calibri', size=11, bold=True)

    fill_bg = PatternFill(fgColor='E8E8E8', fill_type='solid')

    for sheet in book.sheetnames:
        max_width = 0
        current_sheet = book[sheet]
        for i, row in enumerate(current_sheet.iter_rows()):
            for j, cell in enumerate(row):
                if i == 0:
                    cell.font = font_bold
                    cell.fill = fill_bg
                else:
                    cell.font = font

                if i > 0 and j == 0 and len(cell.value) > max_width:
                    max_width = len(cell.value)

        current_sheet.column_dimensions['A'].width = math.ceil(max_width * 1.1)
                
    return book


def generate_filename():
    timestamp = datetime.now().strftime('%y%m%d%H%M%S')
    return f"processed_{timestamp}.xlsx"


def check_auth(username, password):
    return username == AUTH_USERNAME and password == AUTH_PASSWORD


def require_basic_auth(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.authorization
        if not auth or not check_auth(auth.username, auth.password):
            return jsonify({'error': 'Unauthorized'}), 401
        return f(*args, **kwargs)
    return decorated


@app.route('/api/check-auth', methods=['GET'])
@require_basic_auth
def check_auth_route():
    return jsonify({'message': 'OK'}), 200


@app.route('/api/upload', methods=['POST'])
@require_basic_auth
def upload_file():
    file_type = request.form.get('type')  # 'group' o 'arrange'

    if not file_type:
        return jsonify({'error': 'No type specified'}), 400
    
    if 'file' not in request.files:
        return jsonify({'error': 'No file part'}), 400

    file = request.files['file']
    if file.filename == '':
        return jsonify({'error': 'No file selected'}), 400
    
    filename = secure_filename(file.filename)

    if file and allowed_file(file.filename):
        # Leer Excel directamente en memoria sin guardar
        try:
            if file_type == 'group':
                df = pd.read_excel(file, sheet_name=[0], engine='openpyxl')[0].dropna(how='all').dropna(how='all', axis=1)
                headers = df.iloc[0]
                df = pd.DataFrame(df.values[1:], columns=headers)
                return jsonify({
                    'uploaded': filename,
                    'message': 'File uploaded successfully',
                })
            elif file_type == 'arrange':
                # sheet_name=None para leer todas las hojas
                xlsx = pd.read_excel(file, sheet_name=None, engine='openpyxl')
                sheet_names = list(xlsx.keys())            

                df2 = xlsx[sheet_names[1]].dropna(how='all')

                items = df2.columns.to_list()
                
                # Como ejemplo, solo retornamos el nombre original
                return jsonify({
                    'uploaded': filename,
                    'available_columns': items,
                    'message': 'File uploaded successfully'
                })
        except ValueError as e:
            return jsonify({'error': 'Invalid file format', 'details': str(e)}), 400
        except Exception as e:
            return jsonify({'error': 'Read file error', 'details': str(e)}), 400
        finally:
            # Opcional: guardar el archivo en disco si necesitas
            upload_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            # Volver al inicio del archivo porque pandas ya leyó el stream
            file.seek(0)
            file.save(upload_path)

    return jsonify({'error': 'Invalid file type'}), 400


@app.route('/api/process', methods=['POST'])
@require_basic_auth
def process_file():
    data = request.get_json()
    
    file_type = data.get('type')
    
    if not file_type:
        return jsonify({'error': 'No type specified'}), 400

    try:
        if file_type == 'group':
            filename = data['filename']
            name_file_output = process_group(filename)
        elif file_type == 'arrange':
            if 'columns_filter' not in data or 'columns_arrange' not in data:
                return jsonify({'error': 'Faltan columnas para filtrar o arreglar'}), 400
            
            filename = data['filename']
            columns_filter = data['columns_filter']
            columns_arrange = data['columns_arrange']

            name_file_output = process_arrange(filename, columns_filter, columns_arrange)
    except ValueError as e:
        return jsonify({'error': 'Invalid file format', 'details': str(e)}), 400
    except Exception as e:
        return jsonify({'error': 'Failed to process file', 'details': str(e)}), 500
            
    return jsonify({'message': 'File processed successfully', 'filename': name_file_output}), 200


@app.route('/api/download/<filename>', methods=['GET'])
@require_basic_auth
def download_file(filename):
    return send_from_directory(app.config['PROCESSED_FOLDER'], filename, as_attachment=True)


def process_group(filename):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    df = pd.read_excel(file_path, sheet_name=[0], engine='openpyxl')[0].dropna(how='all').dropna(how='all', axis=1)
    headers = df.iloc[0]
    df = pd.DataFrame(df.values[1:], columns=headers)
    
    name_file_output = generate_filename()
    path_file_output = os.path.join(app.config['PROCESSED_FOLDER'], name_file_output)
        
    excel_writer = pd.ExcelWriter(path_file_output, engine='openpyxl')

    # Punto 1
    df_ = df.copy()
    df_['All Science Journal Classification (ASJC) field name'] = df_['All Science Journal Classification (ASJC) field name'].str.split('|')
    df_ = df_.explode('All Science Journal Classification (ASJC) field name')

    table_pivot = pd.crosstab(index=df_['All Science Journal Classification (ASJC) field name'], columns=df_['Year'])
    table_pivot['Total'] = table_pivot.sum(axis=1)
    table_pivot.to_excel(excel_writer, 'by_name_area')

    # Punto 2
    df_ = df.copy()
    df_['All Science Journal Classification (ASJC) field name'] = df_['All Science Journal Classification (ASJC) field name'].apply(lambda x: f"{len(x.split('|'))} areas")

    table_pivot = pd.crosstab(index=df_['All Science Journal Classification (ASJC) field name'], columns=df_['Year'])
    table_pivot['Total'] = table_pivot.sum(axis=1)

    table_pivot.to_excel(excel_writer, 'by_count_area')

    # Punto 3
    df_ = df.copy()
    column = df.filter(like='Sustainable Development Goals', axis=1)
    df_['Sustainable Development Goals'] = column
    df_['Sustainable Development Goals'] = df_['Sustainable Development Goals'].str.split('|').apply(lambda x: [s.strip() for s in x] if isinstance(x, list) else x)
    df_ = df_.explode('Sustainable Development Goals')

    table_pivot = pd.crosstab(index=df_['Sustainable Development Goals'], columns=df_['Year'])
    table_pivot['Total'] = table_pivot.sum(axis=1)

    table_pivot.to_excel(excel_writer, 'by_name_sdg')

    # Punto 4
    df_ = df.copy()
    column = df.filter(like='Sustainable Development Goals', axis=1)
    df_['Sustainable Development Goals'] = column
    df_['Sustainable Development Goals'] = df_['Sustainable Development Goals'].apply(lambda x: f"{0 if x.strip() == '-' else len(x.split('|'))} SDGs")

    table_pivot = pd.crosstab(index=df_['Sustainable Development Goals'], columns=df_['Year'])
    table_pivot['Total'] = table_pivot.sum(axis=1)

    table_pivot.to_excel(excel_writer, 'by_count_sdg')

    set_styless(excel_writer.book)

    # Save file
    excel_writer.book.save(path_file_output)
    
    return name_file_output


def process_arrange(filename, columns_filter, columns_arrange):
    file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
    
    xlsx = pd.read_excel(file_path, sheet_name=None, engine='openpyxl')
    sheet_names = list(xlsx.keys())
    df1 = xlsx[sheet_names[0]].dropna(how='all')
    headers1 = df1.iloc[0]
    df1 = pd.DataFrame(df1.values[1:], columns=headers1)
    df2 = xlsx[sheet_names[1]].dropna(how='all')
    df3 = xlsx[sheet_names[2]].dropna(how='all')
    df3.dropna(subset=['ASJC Code'], inplace=True)
    
    df1_ = df1.loc[df1['Subcategory'] != '-'].copy()
    df1_['Concated'] = df1_['Subject Area'] + '|' + df1['Subcategory']

    list2 = df3['Description'].to_list()

    df1_['sort'] = df1_['Subcategory'].apply(lambda x: list2.index(x) if x in list2 else len(list2))
    df1_ = df1_.sort_values(by='sort').drop('sort', axis=1)

    areas_ = df1_['Subject Area'].unique().tolist()

    fills_ = { i: generate_color() for i in areas_ }

    columns_ = []

    columns_default = ['Scopus Sub-Subject Area']
    [columns_.append(x) if x not in columns_ else set for x in [*columns_filter, *columns_default]]
    df2_ = df2[columns_]

    columns_output = ['#', *columns_]
    df_op = pd.DataFrame([], columns=columns_output)

    len_columns = len(columns_output)
    max_ = np.zeros(len_columns, dtype='intc')

    list1 = df1_['Concated'].unique().tolist()

    for area in list1:
        area_ = area.split('|')[0]
        subarea = area.split('|')[1]
        df_op = pd.concat([df_op, pd.DataFrame([[-1, area_, '', ''], [-2, subarea, '', ''], columns_output], columns=columns_output)])
        sub_set = df2_.loc[df2_['Scopus Sub-Subject Area'] == subarea].copy()
        sub_set.reset_index(drop=True, inplace=True)
        # Arrange columns based on user selection
        columns = [item['value'] for item in columns_arrange]
        ascending = [item['direction'] == 'ascending' for item in columns_arrange]
        sub_set = sub_set.sort_values(by=columns, ascending=ascending)
        sub_set.insert(0, '#', range(1, len(sub_set) + 1))

        widths = pd.DataFrame(sub_set.map(lambda x: len(str(x))).max())
        widths.dropna(how='all', inplace=True)
        widths = widths[0].tolist()
        if len(widths) > 0:
            max_ = [max_[i] if max_[i] > width else width for i, width in enumerate(widths)]
        

        df_op = pd.concat([df_op, sub_set])

    if 'Scopus Sub-Subject Area' not in columns_filter:
        index_ = df_op.columns.tolist().index('Scopus Sub-Subject Area')
        max_.pop(index_)
        df_op.drop(columns=['Scopus Sub-Subject Area'], inplace=True)
        len_columns = len(max_)
    
    name_file_output = generate_filename()
    path_file_output = os.path.join(app.config['PROCESSED_FOLDER'], name_file_output)
    
    excel_writer = pd.ExcelWriter(path_file_output, engine='openpyxl')
    df_op.to_excel(excel_writer, 'results', index=False, header=False)

    sheet = excel_writer.book.active

    font_bold = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
    alignment_ = Alignment(horizontal="center", vertical="center")

    font_bold_sub = Font(name='Calibri', size=18, bold=True, color='843D0D')
    fill_bg_sub = PatternFill(fgColor='DEEBF7', fill_type='solid')
    alignment_sub = Alignment(horizontal="center", vertical="center")

    font_bold_header = Font(name='Calibri', size=11, bold=True, color='FFFFFF')
    fill_bg_header = PatternFill(fgColor='5B9BD5', fill_type='solid')
    alignment_header = Alignment(horizontal="left", vertical="center")

    border = Border(
        left=Side(border_style='thin', color='000000'),
        right=Side(border_style='thin', color='000000'),
        top=Side(border_style='thin', color='000000'),
        bottom=Side(border_style='thin', color='000000')
    )

    for i, row in enumerate(sheet.iter_rows()):
        for j, cell in enumerate(row):
            cell.border = border
            if cell.value == -1:
                sheet.row_dimensions[i + 1].height = 26
                val_ = sheet.cell(i + 1, j + 2).value
                cell.value = f'Scopus Suject Area "{sheet.cell(i + 1, j + 2).value}"'
                letters = [f'{get_column_letter(k)}{str(i + 1)}' for k in range(1, len_columns + 1)]
                sheet.merge_cells(f'{letters[0]}:{letters[-1]}')
                cell.font = font_bold
                cell.fill = PatternFill(fgColor=fills_[val_], fill_type='solid')
                cell.alignment = alignment_
            if cell.value == -2:
                sheet.row_dimensions[i + 1].height = 26
                cell.value = f'Scopus Sub Suject Area "{sheet.cell(i + 1, j + 2).value}"'
                letters = [f'{get_column_letter(k)}{str(i + 1)}' for k in range(1, len_columns + 1)]
                sheet.merge_cells(f'{letters[0]}:{letters[-1]}')
                cell.font = font_bold_sub
                cell.fill = fill_bg_sub
                cell.alignment = alignment_sub
            if cell.value == "#":
                sheet.row_dimensions[i + 1].height = 20
                letters = [f'{get_column_letter(k)}{str(i + 1)}' for k in range(1, len_columns + 1)]
                range_ = sheet[f'{letters[0]}:{letters[-1]}']
                for row_ in range_:
                    for cell_ in row_:
                        cell_.font = font_bold_header
                        cell_.fill = fill_bg_header
                        cell_.alignment = alignment_header

    for i in range(len(max_)):
        letter = get_column_letter(i + 1)
        if max_[i] > 100:
            max_[i] = 75
        sheet.column_dimensions[letter].width = math.ceil(max_[i] * 1.1)

    excel_writer.book.save(path_file_output)

    return name_file_output


@app.route('/')
def index():
    return send_from_directory('frontend', 'index.html')


if __name__ == '__main__':
    app.run(debug=True, host='0.0.0.0', port=5000)
